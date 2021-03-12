from openpyxl import load_workbook
from reporter.utils import open_xls_as_xlsx
from reporter.exceptions import UnsupportedSogazCode
from datetime import datetime


def str_to_date(dates):
    dates = [datetime.strptime(date, "%d.%m.%Y") for date in dates]
    return dates


def sum_list(program):
    s = []
    for i in program:
        s += i
    return s


def get_code(program, file):
    string_program = program
    program = program.split(';')
    program = sum_list([p.split(" ") for p in program])
    program = [p.lower() for p in program]
    program = sum_list([p.split("_") for p in program])
    if {"\"аброссия\"", "аброссия", "аб"} & set(program):
        code = [["АБС СТОМАТОЛОГИЯ Прямой доступ", "ДС№9 к 0618RP137 АБРОССИЯ", 1102]]
    elif "невское" in program:
        code = [["Невское ПКБ СОГАЗ", "ДС№10 к 0618RP137", 4001439]]
    elif "мариинский" in program:
        code = [["Мариинский театр Прямой доступ ПК", "0618RР137 Мариинский театр Прямой доступ ПК", 4001438],
                ["Мариинский театр Прямой доступ Согаз КДЦ", "0618RВ138 Мариинский театр Прямой доступ КДЦ", 4001438]]
    elif "арктика" in program:
        code = [["АО «СОГАЗ» Средний чек ООО СКФ Арктика ПК", "0618RP137 Средний чек Арктика ПК", 4001439],
                ["АО «СОГАЗ» Средний чек ООО СКФ Арктика", "0618RВ138 Средний чек Арктика КДЦ", 4001439]]
    elif {"амб.", "пнд", "амбулаторный"} & set(program):
        code = [["СОГАЗ ПРЯМОЙ ДОСТУП ПК", "0618RР137 ПРЯМОЙ ДОСТУП ПК", 4000971]]
        if "стоматология" in program:
            code += [["СОГАЗ СТОМАТОЛОГИЯ Прямой доступ ПК", "0618RР137 СТОМАТОЛОГИЯ ПРЯМОЙ ДОСТУП ПК", 1102]]
    elif "стоматология" in program:
        code = [["СОГАЗ СТОМАТОЛОГИЯ Прямой доступ ПК", "0618RР137 СТОМАТОЛОГИЯ ПРЯМОЙ ДОСТУП ПК", 1102]]
    else:
        raise UnsupportedSogazCode("Код Согаз не распознан", "Не удалось распознать код в файле {0} в "
                                                             "строке {1}".format(file, string_program))
    return code


def get_data_(ws, file, return_header_row=False):
    data = []
    found = False
    header_row = 0
    for i, row in enumerate(ws):
        row_values = [cell.value for cell in row]
        if found:
            if row[0].value is not None and row[0].value != "":
                data.append([cell.value for cell in row[1:]])
            else:
                break
        elif {"Имя", "Фамилия", "Отчество"}.issubset(set(row_values)):
            found = True
            header_row = i - 1
    if return_header_row:
        return data, header_row
    return data


def get_attach_data(ws, file):
    data, header_row = get_data_(ws, file, True)
    data_formatted = []
    list_type = ws.cell(header_row, 1).value
    list_type = list_type.lower().split(" ")
    for row in data:
        if "изменение" in list_type:
            codes = get_code(row[10], file)
            for code in codes:
                data_formatted.append(row[0:3] + str_to_date([row[3]]) + [row[4]] + str_to_date(row[8:10]) + [None] +
                                      code)
        elif "прикрепление" in list_type:
            codes = get_code(row[12], file)
            for code in codes:
                data_formatted.append(row[0:3] + str_to_date([row[3]]) + [row[9]] + str_to_date(row[10:12]) + [None] +
                                      code)
    return data_formatted


def get_detach_data(ws, file):
    data = get_data_(ws, file)
    data_formatted = []
    for row in data:
        codes = get_code(row[6], file)
        for code in codes:
            data_formatted.append(row[0:3] + str_to_date([row[3]]) + [row[4]] + [None, None] + str_to_date([row[5]]) +
                                  code[0:2] + [2])
    return data_formatted


def create_attach_report(files: list):
    data = []
    for file in files:
        if file.split(".")[-1] == "xls":
            wb_tmp = open_xls_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_attach_data(ws_tmp, file)
    return data


def create_detach_report(files: list):
    data = []
    for file in files:
        if file.split(".")[-1] == "xls":
            wb_tmp = open_xls_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_detach_data(ws_tmp, file)
    return data


if __name__ == '__main__':
    create_attach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Согаз/Прикреп/v8_7CE0_973c.xls",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Согаз/Прикреп/v8_FA1F_7e6a.xls",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Согаз/Прикреп/v8_FA1F_ad07.xls",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Согаз/Прикреп/v8_FA1F_c443.xls"
         ]
    )
    create_detach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Согаз/Откреп/v8_7CE0_8a2f.xls",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Согаз/Откреп/v8_5578_63cc.xls"
        ]
    )
