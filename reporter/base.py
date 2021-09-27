from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from abc import ABC, abstractmethod
from .utils import open_xls_as_xlsx, open_csv_as_xlsx
from . import column_types as ct
from typing import ClassVar, List
from openpyxl.cell import Cell
from functools import singledispatchmethod

ENDING_ROW_CELLS = (
    '',
    None
)


class AbstractReport(ABC):
    def __init__(self, file, encoding=None):
        self.file = file
        self.encoding = encoding
        self.ws = self._get_ws(file)

    @abstractmethod
    def get_data(self):
        pass

    def _get_ws(self, file) -> Worksheet:
        if file.split(".")[-1] in ["xls", "XLS"]:
            wb_tmp = open_xls_as_xlsx(file, self.encoding)
        elif file.split(".")[-1] == "csv":
            wb_tmp = open_csv_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        return ws_tmp

    @staticmethod
    def is_reportable(file):
        return AbstractReport.validate_extension(file)

    @staticmethod
    def validate_extension(file):
        return file.split(".")[-1] in ["xls", "XLS", "csv", "xlsx"]


class Report(AbstractReport):
    """ Отчет """

    def __init__(self, file, config):
        super().__init__(file)
        self.tables = []
        self.typed_tables = []
        self.final_table = []
        self.row_length = 11
        self.codes = config.get("codes")
        self.header_row = config.get("header_row")
        self.ending_row_cells = config.get("ending_row_cells")

    def get_data(self):
        self._get_tables()
        self.typed_tables = [self._make_typed_table(table) for table in self.tables]
        self._make_final_table()
        self._add_codes()
        return self.final_table

    def _make_final_table(self):
        for table in self.typed_tables:
            for row in table:
                final_row = [None] * self.row_length
                for cell in row:
                    self._add_cell_to_row(cell, final_row)
                self.final_table.append(final_row)

    def _add_codes(self):
        new_final_table = []
        for row in self.final_table:
            for code in self.codes:
                copy_row = row.copy()
                self._add_cell_to_row(code, copy_row)
                new_final_table.append(copy_row)
        self.final_table = new_final_table

    @singledispatchmethod
    def _add_cell_to_row(self, cell, row):
        row[cell.column_number] = cell.value

    @_add_cell_to_row.register
    def _(self, cell: ct.FIO, row):
        for sub_value in cell.value:
            row[sub_value.column_number] = sub_value.value

    @_add_cell_to_row.register
    def _(self, cell: ct.Codes, row):
        for sub_value in cell.value:
            row[sub_value.column_number] = sub_value.value

    def _make_typed_table(self, table: List[List[Cell]]) -> List[List[ct.ColumnType]]:
        typed_table = []
        type_map = [
            self._get_column_type(header.value.lower().strip())
            if header.value is not None else None
            for header in table[0]
        ]
        for row in table[1:]:
            typed_row = [type_map[i](cell.value) for i, cell in enumerate(row) if type_map[i] is not None]
            typed_table.append(typed_row)
        return typed_table

    def _get_tables(self):
        """ Метод разделяет таблицу на подтаблицы, отделяя лишнюю информацию """
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.lower() in self.header_row:
                        min_row = cell.row
                        max_row = self._find_max_row(min_row)
                        self.tables.append(self.ws[min_row:max_row])
                        break

    def _find_max_row(self, min_row: int):
        """ Метод для поиска максимального ряда подтаблицы по минимальному """
        for row in self.ws.iter_rows(min_row=min_row):
            if self._is_ending_row(row):
                return row[0].row - 1
        return self.ws.max_row

    def _get_column_type(self, column_header: str) -> ClassVar[ct.ColumnType]:
        return self.header_row.get(column_header)

    def _is_ending_row(self, row):

        for cell in row:
            if cell.value in self.ending_row_cells:
                return True
            else:
                return False


class NullReport(AbstractReport):
    def get_data(self):
        return []

    def _get_ws(self, file):
        return None

