from openpyxl import load_workbook
from reporter.utils import open_xls_as_xlsx, open_csv_as_xlsx
from reporter.config import header, codes, ids
from reporter.exceptions import UnsupportedNameLength, UnrecognisedType

target = ["ФИО", "Номер полиса", "Дата начала обслуживания", "Дата окончания обслуживания", "Программа обслуживания",
          "Дата рождения", "Страхователь", "Дата отмены"]
column_map_attach = {
    "Номер полиса": "POLICY", "Дата начала обслуживания": "DATE_FRM", "Дата окончания обслуживания": "DATE_TO",
    "Дата рождения": "DATE_BIRTH", "Дата отмены": "DATE_CNCL"
}
column_map_detach = {
    "Номер полиса": "POLICY",
    "Дата рождения": "DATE_BIRTH", "Дата окончания обслуживания": "DATE_CNCL"
}


def format_names(names, file_name):
    surname = []
    first_name = []
    sec_name = []
    for i, name in enumerate(names):
        name_array = name.split(" ")
        if len(name_array) == 3:
            surname.append(name_array[0])
            first_name.append(name_array[1])
            sec_name.append(name_array[2])
        elif len(name_array) == 4:
            surname.append(name_array[0] + name_array[1])
            first_name.append(name_array[2])
            sec_name.append(name_array[3])
        else:
            raise UnsupportedNameLength("Ошибка имени", "Неподдерживаемый формат имени '{0}' в {1} строке  в файле {2}"
                                        .format(name, i + 1, file_name))
    return {"SURNAME": surname, "FIRST_NAME": first_name, "SEC_NAME": sec_name}


def get_data(ws, column_map, file_name):
    data = {}
    for row in ws.iter_rows(max_row=1):
        for i, cell in enumerate(row):
            column_name = cell.value
            if column_name in target:
                column = [row[0].value for row in ws.iter_rows(min_row=2, min_col=i + 1, max_col=i + 1)]
                data.update({column_map.get(column_name, column_name): column})
    names = data.pop("ФИО")
    data.update(format_names(names, file_name))
    formatted_data = [[cell] for cell in data.get(header[0])]
    for head in header[1:8]:
        column = data.get(head)
        for i, row in enumerate(formatted_data):
            if column is None:
                row += [None]
            else:
                row += [column[i]]
    return formatted_data


def get_number(file_name):
    number = file_name.split("_")
    if len(number) < 3:
        raise UnrecognisedType("Нераспознанный тип", "Имя файла: {0}".format(file_name))
    number = number[2]
    if number == '1492':
        return "alyans1"
    elif number == '10062':
        return "alyans2"
    else:
        raise UnrecognisedType("Нераспознанный тип", "Найденный тип: {0}\n"
                                                     "Имя файла: {1}".format(number, file_name))


def get_files_data(file, column_map):
    data = []
    if file.split(".")[-1] == "xls":
        wb_tmp = open_xls_as_xlsx(file)
    elif file.split(".")[-1] == "csv":
        wb_tmp = open_csv_as_xlsx(file)
    else:
        wb_tmp = load_workbook(file)
    ws_name = wb_tmp.sheetnames[0]
    ws_tmp = wb_tmp[ws_name]
    data += get_data(ws_tmp, column_map, file.split("/")[-1])
    return data


def create_attach_report(files: list):
    data = []
    for file in files:
        number = get_number(file.split("/")[-1])
        data1 = get_files_data(file, column_map_attach)
        for i, r in enumerate(data1):
            data1[i] = r + codes[number] + ids["alyans_attach"]
        data += data1
    return data


def create_detach_report(files: list):
    data = []
    for file in files:
        number = get_number(file.split("/")[-1])
        data1 = get_files_data(file, column_map_detach)
        for i, r in enumerate(data1):
            data1[i] = r + codes[number] + ids["alyans_detach"]
        data += data1
    return data


if __name__ == '__main__':
    data = create_attach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/??_2021.02.08_1492_???????????.csv",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/??_2021.02.08_10062_???????????.csv"
        ]
    )
