from reporter.base import AbstractReport, NullReport
import reporter.column_types as ct
from typing import ClassVar, List
from openpyxl.cell import Cell
from functools import singledispatchmethod

ATTACH_HEADER_ROW = {
    "фио": ct.FIO,
    "дата рождения": ct.BirthDay,
    "№ полиса": ct.Policy,
    "дата действия полиса с": ct.DateFrom,
    "дата действия полиса по": ct.DateTo,
    "дата действия полиса": ct.DateFrom,
    "дата действия полиса до": ct.DateTo
}

DETACH_HEADER_ROW = {
    "фио": ct.FIO,
    "дата рождения": ct.BirthDay,
    "№ полиса": ct.Policy,
    "дата открепления с (с данной даты не обслуживается)": ct.DateCancel
}

CODES_KDC_ATT = ct.Codes(
                clinic_code=ct.ClinicCode(value='САО "ВСК" КДЦ'),
                control_code=ct.ControlCode(value='ДС10к045СК/2016/16180SMU00009'),
                medicine_id=ct.MedicinesID(value=4001554)
            )
CODES_PK_ATT = ct.Codes(
                clinic_code=ct.ClinicCode(value='САО "ВСК" ПК'),
                control_code=ct.ControlCode(value='ДС19к17180SMU00092/9020'),
                medicine_id=ct.MedicinesID(value=4001554)
            )


CODES_KDC_DET = ct.Codes(
                clinic_code=ct.ClinicCode(value='САО "ВСК" КДЦ'),
                control_code=ct.ControlCode(value='ДС10к045СК/2016/16180SMU00009'),
                medicine_id=ct.MedicinesID(value=2)
            )
CODES_PK_DET = ct.Codes(
                clinic_code=ct.ClinicCode(value='САО "ВСК" ПК'),
                control_code=ct.ControlCode(value='ДС19к17180SMU00092/9020'),
                medicine_id=ct.MedicinesID(value=2)
            )

ENDING_ROW_CELLS = (
    '',
    None
)


class AlfaReport(AbstractReport):
    """ Отчет ВСК """

    def __init__(self, file, header_row: dict, codes: List[ct.Codes]):
        super(AlfaReport, self).__init__(file)
        self.tables = []
        self.typed_tables = []
        self.final_table = []
        self.row_length = 11
        self.codes = codes
        self.header_row = header_row

    def get_data(self):
        self._get_tables()
        self.typed_tables = [self._make_typed_table(table) for table in self.tables]
        self._make_final_table()
        self._add_codes()
        return self.final_table

    def _make_final_table(self):
        for table in self.typed_tables:
            for row in table:
                final_row = [None] * self.row_length
                for cell in row:
                    self._add_cell_to_row(cell, final_row)
                self.final_table.append(final_row)

    def _add_codes(self):
        new_final_table = []
        for row in self.final_table:
            for code in self.codes:
                copy_row = row.copy()
                self._add_cell_to_row(code, copy_row)
                new_final_table.append(copy_row)
        self.final_table = new_final_table

    @singledispatchmethod
    def _add_cell_to_row(self, cell, row):
        row[cell.column_number] = cell.value

    @_add_cell_to_row.register
    def _(self, cell: ct.FIO, row):
        for sub_value in cell.value:
            row[sub_value.column_number] = sub_value.value

    @_add_cell_to_row.register
    def _(self, cell: ct.Codes, row):
        for sub_value in cell.value:
            row[sub_value.column_number] = sub_value.value

    def _make_typed_table(self, table: List[List[Cell]]) -> List[List[ct.ColumnType]]:
        typed_table = []
        type_map = [
            self._get_column_type(header.value.lower().strip())
            if header.value is not None else None
            for header in table[0]
        ]
        for row in table[1:]:
            typed_row = [type_map[i](cell.value) for i, cell in enumerate(row) if type_map[i] is not None]
            typed_table.append(typed_row)
        return typed_table

    def _get_tables(self):
        """ Метод разделяет таблицу на подтаблицы, отделяя лишнюю информацию """
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.lower() in self.header_row:
                        min_row = cell.row
                        max_row = self._find_max_row(min_row)
                        self.tables.append(self.ws[min_row:max_row])
                        break

    def _find_max_row(self, min_row: int):
        """ Метод для поиска максимального ряда подтаблицы по минимальному """
        for row in self.ws.iter_rows(min_row=min_row):
            if self._is_ending_row(row):
                return row[0].row - 1
        return self.ws.max_row

    def _get_column_type(self, column_header: str) -> ClassVar[ct.ColumnType]:
        return self.header_row.get(column_header)

    @staticmethod
    def _is_ending_row(row):

        for cell in row:
            if cell.value in ENDING_ROW_CELLS:
                return True
            else:
                return False


def get_alfa(file, attach=False, detach=False):
    if AlfaReport.is_reportable(file):
        if attach:
            return AlfaReport(file, ATTACH_HEADER_ROW, [CODES_KDC_ATT, CODES_PK_ATT])
        elif detach:
            return AlfaReport(file, DETACH_HEADER_ROW, [CODES_KDC_DET, CODES_PK_DET])
    else:
        return NullReport(file)
