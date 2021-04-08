from openpyxl import load_workbook
from reporter.utils import open_xls_as_xlsx
from reporter.config import codes, ids
from reporter.exceptions import UnsupportedRenCode
import re
from reporter.base import BaseInsurance
import datetime

target = ["Фамилия", "Имя", "Отчество", "Дата рождения", "Номер полиса"]


def get_date(ws):
    text = ws.cell(14, 1).value
    date = re.search('([0-2]\d|3[01]).(0\d|1[012]).(\d{4})', text).group()
    date = datetime.datetime.strptime(date, '%d.%m.%Y')
    return date


class Renessans(BaseInsurance):
    def __init__(self):
        super().__init__()
        self.detach_date = None
        self.detach_code = None

    def get_attach(self, ws):
        data = []
        for row in ws.iter_rows(min_row=3, max_col=7):
            data += self.parse_attach_row(self.parse_attach_row(row))
        return data

    @staticmethod
    def parse_attach_row(row):
        r = [cell.value for cell in row]
        r.insert(8, None)
        return [r + codes["renessans2"] + ids["renessans_attach"],
                r + codes["renessans1"] + ids["renessans_attach"]]

    def get_detach(self, ws):
        data = []
        found = False
        self.detach_date = get_date(ws)
        self.detach_code = self.get_detach_code(ws)
        for row in ws:
            if found:
                if row[0].value is not None and row[0].value != "":
                    data.append(self.parse_detach_row(row))
                else:
                    return data
            for i, cell in enumerate(row[1:6]):
                if cell.value == target[i]:
                    if not found:
                        found = True

    def parse_detach_row(self, row):
        r = [cell.value for cell in row[1:6]] + [self.detach_date] + self.detach_code + ids["renessans_detach"]
        r.insert(5, None)
        r.insert(5, None)
        return r

    def get_detach_code(self, ws):
        detach_type = ws.cell(1, 5).value
        if detach_type == "АО \"Поликлинический комплекс\"":
            return codes["renessans1"]
        elif detach_type == "АО \"Современные медицинские технологии\"":
            return codes["renessans2"]
        else:
            raise UnsupportedRenCode("Ошибка кода Ренессанс", "Неподдерживаемая комания {0}, \n "
                                                              "В файле {1}".format(detach_type, self.file))


def get_detach_code(ws, file):
    detach_type = ws.cell(1, 5).value
    if detach_type == "АО \"Поликлинический комплекс\"":
        return codes["renessans1"]
    elif detach_type == "АО \"Современные медицинские технологии\"":
        return codes["renessans2"]
    else:
        raise UnsupportedRenCode("Ошибка кода Ренессанс", "Неподдерживаемая комания {0}, \n "
                                                          "В файле {1}".format(detach_type, file))


def get_date(ws):
    text = ws.cell(14, 1).value
    date = re.search('([0-2]\d|3[01]).(0\d|1[012]).(\d{4})', text).group()
    date = datetime.datetime.strptime(date, '%d.%m.%Y')
    return date


def get_detach_data(ws, file):
    data = []
    found = False
    date = get_date(ws)
    detach_code = get_detach_code(ws, file)
    for row in ws:
        if found:
            if row[0].value is not None and row[0].value != "":
                data.append([cell.value for cell in row[1:6]] + [date] + detach_code + ids["renessans_detach"])
            else:
                return data
        for i, cell in enumerate(row[1:6]):
            if cell.value == target[i]:
                if not found:
                    found = True


def get_attach_data(ws):
    data = []
    for row in ws.iter_rows(min_row=3, max_col=7):
        data.append([cell.value for cell in row])
    return data


def create_detach_report(files: list):
    data = []
    for file in files:
        if file.split(".")[-1] == "xls":
            wb_tmp = open_xls_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_detach_data(ws_tmp, file)
    for r in data:
        r.insert(5, None)
        r.insert(5, None)
    return data


def create_attach_report(files: list):
    data = []
    for file in files:
        if file.split(".")[-1] == "xls":
            wb_tmp = open_xls_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_attach_data(ws_tmp)
    for r in data:
        r.insert(8, None)
    data1 = [r + codes["renessans2"] + ids["renessans_attach"] for r in data]
    for i, r in enumerate(data):
        data[i] = r + codes["renessans1"] + ids["renessans_attach"]
    data += data1
    return data


if __name__ == '__main__':
    detach = Renessans()
    detach.create_detach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Ренессанс Откреп/002_1060_001???35719520?_19_????_03_02_2021.xls",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Ренессанс Откреп/002_1060_001???35890520?_3_????_01_02_2021.xls"
        ]
    )
    attach = Renessans()
    attach.create_attach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Ренессанс Прикреп/002_1682_001???35682620?_79_?????_27_01_2021 (1).xls"
        ]
    )
