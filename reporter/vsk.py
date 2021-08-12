from reporter.base import Report, NullReport
import reporter.column_types as ct
from typing import ClassVar, List, Dict
from openpyxl.cell import Cell

ATTACH_HEADER_ROW = {
    "фио": ct.FIO,
    "др": ct.BirthDay,
    "полис": ct.Policy,
    "дата действия полиса с": ct.DateFrom,
    "дата действия полиса по": ct.DateTo
}

DETACH_HEADER_ROW = {
    "фио": ct.FIO,
    "др": ct.BirthDay,
    "полис": ct.Policy,
    "дата действия полиса по": ct.DateCancel
}

ENDING_ROW_CELLS = (
    '',
    None
)


class VSKReport(Report):
    """ Отчет ВСК """

    def __init__(self, file, header_row: Dict):
        super(VSKReport, self).__init__(file)
        self.tables = []
        self.typed_tables = []
        self.final_table = []
        self.row_length = 11
        self.header_row = header_row

    def get_data(self):
        self._get_tables()
        self.typed_tables = [self._make_typed_table(table) for table in self.tables]
        self._make_final_table()
        return self.final_table

    def _make_final_table(self):
        for table in self.typed_tables:
            for row in table:
                final_row = [None] * self.row_length
                for cell in row:
                    if isinstance(cell, ct.FIO):
                        for sub_value in cell.value:
                            final_row[sub_value.column_number] = sub_value.value
                    else:
                        final_row[cell.column_number] = cell.value
                self.final_table.append(final_row)

    def _make_typed_table(self, table: List[List[Cell]]) -> List[List[ct.ColumnType]]:
        typed_table = []
        type_map = [self._get_column_type(header.value.lower().strip()) for header in table[0]]
        for row in table[1:]:
            typed_row = [type_map[i](cell.value) for i, cell in enumerate(row) if type_map[i] is not None]
            typed_table.append(typed_row)
        return typed_table

    def _get_tables(self):
        """ Метод разделяет таблицу на подтаблицы, отделяя лишнюю информацию """
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.lower() in self.header_row:
                        min_row = cell.row
                        max_row = self._find_max_row(min_row)
                        self.tables.append(self.ws[min_row:max_row])
                        break

    def _find_max_row(self, min_row: int):
        """ Метод для поиска максимального ряда подтаблицы по минимальному """
        for row in self.ws.iter_rows(min_row=min_row):
            if self._is_ending_row(row):
                return row[0].row - 1
        return self.ws.max_row

    def _get_column_type(self, column_header: str) -> ClassVar[ct.ColumnType]:
        return self.header_row.get(column_header)

    @staticmethod
    def _is_ending_row(row):

        for cell in row:
            if cell.value in ENDING_ROW_CELLS:
                return True
            else:
                return False


def get_vsk(file, attach=False, detach=False):
    if VSKReport.is_reportable(file):
        if attach:
            return VSKReport(file, ATTACH_HEADER_ROW)
        elif detach:
            return VSKReport(file, DETACH_HEADER_ROW)
    else:
        return NullReport(file)
