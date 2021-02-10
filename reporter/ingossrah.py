from openpyxl import load_workbook, Workbook
from reporter.utils import open_xls_as_xlsx, open_csv_as_xlsx, fit_columns_width
import os
from reporter.config import header, codes, ids
from reporter.exceptions import UnsupportedNameLength

target = ["ФИО", "Номер полиса", "Дата начала обслуживания", "Дата окончания обслуживания", "Программа обслуживания",
          "Дата рождения", "Страхователь", "Дата отмены"]
column_map_attach = {
    "Номер полиса": "POLICY", "Дата начала обслуживания": "DATE_FRM", "Дата окончания обслуживания": "DATE_TO",
    "Дата рождения": "DATE_BIRTH", "Дата отмены": "DATE_CNCL"
}
column_map_detach = {
    "Номер полиса": "POLICY",
    "Дата рождения": "DATE_BIRTH", "Дата окончания обслуживания": "DATE_CNCL"
}
column_map_attach_fullrisk = {
    "Полис": "POLICY",
    "Фамилия": "SURNAME",
    "Имя": "FIRST_NAME",
    "Отчество": "SEC_NAME",
    "Д.Рожд.": "DATE_BIRTH",
    "Дата прикрепления": "DATE_FRM",
    "Дата открепления": "DATE_TO"
}

column_map_attach_kdp4 = {
    "Фамилия": "SURNAME",
    "Имя": "FIRST_NAME",
    "Отчество": "SEC_NAME",
    "Дата рождения": "DATE_BIRTH",
    "№ полиса": "POLICY",
    "Начало обслуживания": "DATE_FRM",
    "Конец обслуживания": "DATE_TO"
}

column_map_detach_fullrisk = {
    "Полис": "POLICY",
    "Фамилия": "SURNAME",
    "Имя": "FIRST_NAME",
    "Отчество": "SEC_NAME",
    "Д.Рожд.": "DATE_BIRTH",
    "Дата открепления": "DATE_CNCL"
}

column_map_detach_kdp4 = {
    "Фамилия": "SURNAME",
    "Имя": "FIRST_NAME",
    "Отчество": "SEC_NAME",
    "Дата рождения": "DATE_BIRTH",
    "№ полиса": "POLICY",
    "Конец обслуживания": "DATE_CNCL"
}


def format_data(data):
    formatted_data = [[cell] for cell in data.get(header[0])]
    for head in header[1:8]:
        for i, row in enumerate(formatted_data):
            column = data.get(head)
            if column is not None:
                row += [column[i]]
            else:
                row += [None]
    return formatted_data


def get_max_fullrisk(ws):
    max_row = ws.max_row
    for i, row in enumerate(ws.iter_rows(min_row=14)):
        row = [cell.value for cell in row]
        for y, cell in enumerate(row):
            if cell in ["", None]:
                row[y] = True
            else:
                row[y] = False
        if all(row):
            max_row = 13 + i
            break
    return max_row


def get_data_fullrisk(ws, column_map):
    data = {}
    max_row = get_max_fullrisk(ws)
    for row in ws.iter_rows(min_row=13, max_row=13):
        for cell in row:
            if cell.value in column_map.keys():
                column = []
                for column_row in ws.iter_rows(min_row=14, min_col=cell.column, max_col=cell.column, max_row=max_row):
                    column.append(column_row[0].value)
                data.update({column_map[cell.value]: column})
    return format_data(data)


def get_data_kdp4(ws, column_map):
    data = {}
    for row in ws.iter_rows(min_row=10, max_row=10):
        for cell in row:
            if cell.value in column_map.keys():
                column = []
                for column_row in ws.iter_rows(min_row=11, min_col=cell.column, max_col=cell.column):
                    column.append(column_row[0].value)
                data.update({column_map[cell.value]: column})
    return format_data(data)


def get_files_data(file, attach=False):
    data = []
    if file.split(".")[-1] in ["xls", "XLS"]:
        wb_tmp = open_xls_as_xlsx(file)
    elif file.split(".")[-1] == "csv":
        wb_tmp = open_csv_as_xlsx(file)
    else:
        wb_tmp = load_workbook(file)
    ws_name = wb_tmp.sheetnames[0]
    ws_tmp = wb_tmp[ws_name]
    file_type = file.split(".")[0].split("_")[-1]
    if file_type == "FULLRISK":
        if attach:
            data += get_data_fullrisk(ws_tmp, column_map_attach_fullrisk)
        else:
            data += get_data_fullrisk(ws_tmp, column_map_detach_fullrisk)
    elif file_type == "KDP4":
        if attach:
            data += get_data_kdp4(ws_tmp, column_map_attach_kdp4)
        else:
            data += get_data_kdp4(ws_tmp, column_map_detach_kdp4)
    else:
        raise ValueError
    return data


def create_attach_report(files: list, save_directory: str, filename="FileIngosstrahAttach"):
    wb = Workbook()
    ws = wb.active
    data = [header]
    for file in files:
        data += get_files_data(file, True)
    for i, row in enumerate(data):
        if i == 0:
            ws.append(row)
        else:
            ws.append(row + codes["ingosstrah"] + ids["ingosstrah_attach"])
    fit_columns_width(ws)
    wb.save(os.path.join(save_directory, filename + ".xlsx"))


def create_detach_report(files: list, save_directory: str, filename="FileIngosstrahDetach"):
    wb = Workbook()
    ws = wb.active
    data = [header]
    for file in files:
        data += get_files_data(file)
    for i, row in enumerate(data):
        if i == 0:
            ws.append(row)
        else:
            ws.append(row + codes["ingosstrah"] + ids["ingosstrah_detach"])
    fit_columns_width(ws)
    wb.save(os.path.join(save_directory, filename + ".xlsx"))


if __name__ == '__main__':
    create_attach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/_04_02_2021_11_17_55_111872_FULLRISK.XLS",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/_04_02_2021_11_34_33_978666_KDP4.XLS"
        ],
        "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test")
    create_detach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/_04_02_2021_11_26_38_978666_KDP4.XLS"
        ],
        "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test"
    )
