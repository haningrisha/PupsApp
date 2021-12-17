from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import MergedCell
from abc import ABC, abstractmethod
from .utils import open_xls_as_xlsx, open_csv_as_xlsx
from . import column_types as ct
from typing import ClassVar, List
from openpyxl.cell import Cell
from functools import singledispatchmethod, cached_property

ENDING_ROW_CELLS = (
    '',
    None
)


class Config:
    def __init__(self, config: dict):
        self.config = config

    @cached_property
    def depth(self):
        def count_depth(dictionary):
            depth = 1
            for value in dictionary.values():
                if isinstance(value, dict):
                    if depth < 1 + count_depth(value):
                        depth = 1 + count_depth(value)
            return depth
        return count_depth(self.config)

    def get(self, *args, **kwargs):
        return self.config.get(*args, **kwargs)


class AbstractReport(ABC):
    def __init__(self, file, encoding=None):
        self.file = file
        self.encoding = encoding
        self.ws = self._get_ws(file)

    @abstractmethod
    def get_data(self):
        pass

    def _get_ws(self, file) -> Worksheet:
        if file.split(".")[-1] in ["xls", "XLS"]:
            wb_tmp = open_xls_as_xlsx(file, self.encoding)
        elif file.split(".")[-1] == "csv":
            wb_tmp = open_csv_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        return ws_tmp

    @staticmethod
    def is_reportable(file):
        return AbstractReport.validate_extension(file)

    @staticmethod
    def validate_extension(file):
        return file.split(".")[-1] in ["xls", "XLS", "csv", "xlsx"]


class Report(AbstractReport):
    """ Отчет """

    def __init__(self, file, config: Config):
        super().__init__(file)
        self.tables = []
        self.typed_tables = []
        self.final_table = []
        self.row_length = 11
        self.codes = config.get("codes")
        self.header_row = config.get("header_row")
        self.ending_row_cells = config.get("ending_row_cells")
        self.is_code_filtered = config.get("is_code_filtered", False)
        self.config = config

    def get_data(self):
        self._get_tables()
        self.typed_tables = [self._make_typed_table(table) for table in self.tables]
        self._make_final_table()
        self._add_codes()
        return self.final_table

    def _make_final_table(self):
        for table in self.typed_tables:
            for row in table:
                final_row = [None] * self.row_length
                for cell in row:
                    self._add_cell_to_row(cell, final_row)
                self.final_table.append(final_row)

    def _add_codes(self):
        if not self.is_code_filtered:
            new_final_table = []
            for row in self.final_table:
                for code in self.codes:
                    copy_row = row.copy()
                    self._add_cell_to_row(code, copy_row)
                    new_final_table.append(copy_row)
            self.final_table = new_final_table

    @singledispatchmethod
    def _add_cell_to_row(self, cell, row):
        row[cell.column_number] = cell.value

    @_add_cell_to_row.register
    def _(self, cell: ct.FIO, row):
        for sub_value in cell.value:
            row[sub_value.column_number] = sub_value.value

    @_add_cell_to_row.register
    def _(self, cell: ct.Codes, row):
        for sub_value in cell.value:
            row[sub_value.column_number] = sub_value.value

    @_add_cell_to_row.register
    def _(self, cell: ct.CodeFilter, row):
        try:
            for codes in cell.get_codes():
                for code in codes.value:
                    row[code.column_number] = code.value
        except TypeError:
            pass

    def _make_typed_table(self, table: List[List[Cell]]) -> List[List[ct.ColumnType]]:
        typed_table = []
        type_map = self._get_type_map(table)
        for row in table[self.config.depth-1:]:
            typed_row = [type_map[i](cell.value) for i, cell in enumerate(row) if type_map[i] is not None]
            typed_table.append(typed_row)
        return typed_table

    def _get_type_map(self, table: List[List[Cell]], depth=0, config_header_row=None, col_start=None, col_end=None) -> list:
        type_map = []
        if config_header_row is None:
            config_header_row = self.header_row
        header_row = table[depth]
        if col_start is not None:
            header_row = header_row[col_start:]
        if col_end is not None:
            header_row = header_row[:col_end]
        for header in header_row:
            if header.value is not None:
                column_type = self._get_column_type(config_header_row, header.value.lower().strip())
                if isinstance(column_type, dict):
                    merged_cells = self.count_merged_cells(table[depth], header.column)
                    type_map.extend(
                        self._get_type_map(
                            table, depth+1, column_type, col_start=header.column-1, col_end=merged_cells+header.column
                        )
                    )
                else:
                    type_map.append(column_type)
            elif not isinstance(header, MergedCell):
                type_map.append(None)
        return type_map

    @staticmethod
    def count_merged_cells(row, start=0):
        count = 0
        for cell in row[start:]:
            if isinstance(cell, MergedCell):
                count += 1
            else:
                return count
        return count

    def _get_tables(self):
        """ Метод разделяет таблицу на подтаблицы, отделяя лишнюю информацию """
        for row in self.ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.lower() in self.header_row:
                        min_row = cell.row
                        min_col = cell.column
                        max_row = self._find_max_row(min_row, min_col)
                        self.tables.append(self.ws[min_row:max_row])
                        break

    def _find_max_row(self, min_row: int, min_col: int):
        """ Метод для поиска максимального ряда подтаблицы по минимальному """
        for row in self.ws.iter_rows(min_row=min_row+self.config.depth-1, min_col=min_col):
            if self._is_ending_row(row):
                return row[0].row - 1
        return self.ws.max_row

    @staticmethod
    def _get_column_type(header_row, column_header: str) -> ClassVar[ct.ColumnType]:
        return header_row.get(column_header)

    def _is_ending_row(self, row):
        for cell in row:
            if cell.value in self.ending_row_cells:
                return True
            else:
                return False


class NullReport(AbstractReport):
    def get_data(self):
        return []

    def _get_ws(self, file):
        return None
