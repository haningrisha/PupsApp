from openpyxl import load_workbook
from reporter.utils import open_xls_as_xlsx, open_csv_as_xlsx


def get_attach(ws, file):
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'судебный' in row[1].lower().split(" "):
            codes = [["МАКС КДЦ Судебный департамент", "0034/СК Судебный департамент", 1102],
                     ["МАКС ПК Судебный департамент", "(МАКС) 101725/16-57937/330/2017 Судебный департамент", 1102]]
        else:
            codes = [["МАКС КДЦ", "0034/СК", 1102],
                     ["Макс стандарт", "(МАКС) 101725/16-57937/330/2017", 1102]]
        for code in codes:
            data += [[row[4], row[5], row[6], row[8], row[3], row[9], row[10], None] + code]
    return data


def get_detach(ws, file):
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'судебный' in row[1].lower().split(" "):
            codes = [["МАКС КДЦ Судебный департамент", "0034/СК Судебный департамент", 2],
                     ["МАКС ПК Судебный департамент", "(МАКС) 101725/16-57937/330/2017 Судебный департамент", 2]]
        else:
            codes = [["МАКС КДЦ", "0034/СК", 1102],
                     ["Макс стандарт", "(МАКС) 101725/16-57937/330/2017", 1102]]
        for code in codes:
            data += [[row[4], row[5], row[6], row[8], row[3], None, None, row[10]] + code]
    return data


def create_attach_report(files: list):
    data = []
    for file in files:
        if file.split(".")[-1] in ["xls", "XLS"]:
            wb_tmp = open_xls_as_xlsx(file, "windows-1251")
        elif file.split(".")[-1] == "csv":
            wb_tmp = open_csv_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_attach(ws_tmp, file)
    return data


def create_detach_report(files: list):
    data = []
    for file in files:
        if file.split(".")[-1] in ["xls", "XLS"]:
            wb_tmp = open_xls_as_xlsx(file, "windows-1251")
        elif file.split(".")[-1] == "csv":
            wb_tmp = open_csv_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_detach(ws_tmp, file)
    return data


if __name__ == '__main__':
    create_attach_report(["/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Макс/0501p210315_05322_210312.xls"])
    create_detach_report(["/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/Макс/0501p210315_05322_210312.xls"])
