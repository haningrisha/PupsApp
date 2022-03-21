import xlrd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import csv


def open_xls_as_xlsx(filename, encoding=None):
    book = xlrd.open_workbook(filename, encoding_override=encoding)
    index = 0
    sheet = book.sheet_by_index(index)
    book1 = Workbook()
    sheet1 = book1.active
    for row in sheet:
        data_row = []
        for cell in row:
            if cell.ctype == 3:
                data_row.append(xlrd.xldate_as_datetime(cell.value, 0))
            else:
                data_row.append(cell.value)
        sheet1.append(data_row)
    return book1


def open_csv_as_xlsx(filename):
    book = Workbook()
    sheet = book.active
    with open(filename, newline='') as csvfile:
        reader = csv.reader(csvfile, delimiter=';')
        for row in reader:
            sheet.append(row)
    return book


def fit_columns_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def format_date(ws, column):
    for row in ws.iter_rows(min_row=1, min_col=column, max_col=column):
        row[0].number_format = 'dd.mm.yyyy'


def is_value_in_sheet(ws: Worksheet, value: str) -> bool:
    """ Функция определяет содержит ли таблица значение """
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if not value.lower() in str(cell).strip().lower():
                continue
            return True
    return False


def combine_codes(code1: dict, code2: dict):
    result = {}
    for key in code1.keys():
        result.update({key: code1[key] + code2.get(key, [])})
    return result
