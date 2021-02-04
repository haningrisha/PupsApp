from openpyxl import load_workbook, Workbook
import xlrd
import os
import re
import datetime
target = ["Фамилия", "Имя", "Отчество", "Дата рождения", "Номер полиса"]
detach_header = ["Фамилия", "Имя", "Отчество", "Дата рождения", "Номер полиса", "Дата открепления"]
attach_header = \
    [["SURNAME", "FIRST_NAME", "SEC_NAME", "DATE_BIRTH", "POLICY", "DATE_FRM", "DATE_TO", "ADRES", "TEL", "MP",
      "WORK_PLACE"],
     ["ФАМИЛИЯ", "ИМЯ", "ОТЧЕСТВО", "ДАТ. РОЖД.", "№ ПОЛИСА", "НАЧАЛО", "КОНЕЦ", "АДРЕС", "ТЕЛЕФОН", "МЕД. ПРОГРАММА",
      "МЕСТО РАБОТЫ"]]


def open_xls_as_xlsx(filename):
    book = xlrd.open_workbook(filename)
    index = 0
    sheet = book.sheet_by_index(index)
    book1 = Workbook()
    sheet1 = book1.active
    for row in sheet:
        data_row = []
        for cell in row:
            if cell.ctype == 3:
                data_row.append(xlrd.xldate_as_datetime(cell.value, 0))
            else:
                data_row.append(cell.value)
        sheet1.append(data_row)
    return book1


def fit_columns_width(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value


def get_date(ws):
    text = ws.cell(14, 1).value
    date = re.search('([0-2]\d|3[01]).(0\d|1[012]).(\d{4})', text).group()
    date = datetime.datetime.strptime(date, '%d.%m.%Y')
    return date


def get_data(ws):
    data = []
    found = False
    date = get_date(ws)
    for row in ws:
        if found:
            if row[0].value is not None and row[0].value != "":
                data.append([cell.value for cell in row[1:6]] + [date])
            else:
                return data
        for i, cell in enumerate(row[1:6]):
            if cell.value == target[i]:
                if not found:
                    found = True


def get_attach_data(ws):
    data = []
    for row in ws.iter_rows(min_row=3):
        data.append([cell.value for cell in row])
    return data


def create_detach_report(files: list, save_directory: str, filename="File"):
    wb = Workbook()
    ws = wb.active
    data = [detach_header]
    for file in files:
        if file.split(".")[-1] == "xls":
            wb_tmp = open_xls_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_data(ws_tmp)
    for r in data:
        ws.append(r)
    fit_columns_width(ws)
    wb.save(os.path.join(save_directory, filename+".xlsx"))


def create_attach_report(files: list, save_directory: str, filename="FileAttach"):
    wb = Workbook()
    ws = wb.active
    data = attach_header
    for file in files:
        if file.split(".")[-1] == "xls":
            wb_tmp = open_xls_as_xlsx(file)
        else:
            wb_tmp = load_workbook(file)
        ws_name = wb_tmp.sheetnames[0]
        ws_tmp = wb_tmp[ws_name]
        data += get_attach_data(ws_tmp)
    for r in data:
        ws.append(r)
    fit_columns_width(ws)
    wb.save(os.path.join(save_directory, filename + ".xlsx"))


if __name__ == '__main__':
    create_detach_report(
        [
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/002_1060_001???35890520?_3_????_01_02_2021.xls",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/002_1060_001???35719520?_19_????_03_02_2021.xlsx"
            ],
        "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test")
    create_attach_report(
        [
            "/Users/grigorijhanin/Documents/ТеорМех/динамика пупс/angle.png",
            "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test/002_1682_001???35682620?_79_?????_27_01_2021 (1).xls"
        ],
        "/Users/grigorijhanin/Documents/Работа пупс/PupsApp/test"
    )
